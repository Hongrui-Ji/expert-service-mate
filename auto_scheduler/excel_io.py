from __future__ import annotations

import re
import zipfile
from datetime import date, datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from .engine import ScheduleRunResult

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_INPUT_ROWS = 20_000
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 5_000
DATE_COLUMNS = {"开始日期", "截止日期", "服务锚点日期", "建议开始日期", "建议截止日期"}


def read_input_workbook(content: bytes) -> pd.DataFrame:
    if not content:
        raise ValueError("上传的 Excel 文件为空")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError("Excel 文件不能超过 10 MB")
    archive = BytesIO(content)
    if not zipfile.is_zipfile(archive):
        raise ValueError("上传内容不是有效的 .xlsx 文件")
    archive.seek(0)
    with zipfile.ZipFile(archive) as workbook_archive:
        entries = workbook_archive.infolist()
        if len(entries) > MAX_ARCHIVE_ENTRIES:
            raise ValueError("Excel 文件内部条目过多，已拒绝处理")
        if sum(entry.file_size for entry in entries) > MAX_UNCOMPRESSED_BYTES:
            raise ValueError("Excel 解压后超过 100 MB，已拒绝处理")
    try:
        dataframe = pd.read_excel(BytesIO(content), sheet_name=0, engine="openpyxl")
    except Exception as exc:
        raise ValueError(f"Excel 文件无法读取：{exc}") from exc
    if len(dataframe) > MAX_INPUT_ROWS:
        raise ValueError(f"Excel 数据不能超过 {MAX_INPUT_ROWS:,} 行")
    return dataframe


def _safe_sheet_name(value: Any, used: set[str]) -> str:
    name = re.sub(r"[\[\]:*?/\\]", "_", str(value)).strip() or "未命名"
    name = name[:31]
    base = name
    suffix_number = 1
    while name in used:
        suffix = f"_{suffix_number}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        suffix_number += 1
    used.add(name)
    return name


def _escape_formula(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    safe = dataframe.copy()
    for column in safe.columns:
        if safe[column].dtype == "object":
            safe[column] = safe[column].map(_escape_formula)
    return safe


def _write_dataframe(writer, dataframe: pd.DataFrame, sheet_name: str):
    _safe_dataframe(dataframe).to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    header_names = [cell.value for cell in worksheet[1]]
    for index, name in enumerate(header_names, start=1):
        width = min(42, max(12, len(str(name or "")) * 2 + 2))
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width
        if name in DATE_COLUMNS:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=index)
                if isinstance(cell.value, (date, datetime)):
                    cell.number_format = "yyyy/mm/dd"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def build_output_workbook(result: ScheduleRunResult, *, split_by_province: bool = True) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        used_names: set[str] = set()
        summary_frame = pd.DataFrame(
            [{"项目": key, "数值": value} for key, value in result.summary.items()]
        )
        summary_sheet = _safe_sheet_name("运行摘要", used_names)
        _write_dataframe(writer, summary_frame, summary_sheet)

        scheduled = result.scheduled
        if scheduled.empty:
            result_sheet = _safe_sheet_name("排班结果", used_names)
            _write_dataframe(writer, pd.DataFrame(columns=["提示"], data=[["没有生成可排班任务"]]), result_sheet)
        elif split_by_province and "省份" in scheduled.columns:
            for province, province_frame in scheduled.groupby("省份", dropna=False, sort=True):
                province_name = province if pd.notna(province) and str(province).strip() else "未标注省份"
                sheet_name = _safe_sheet_name(province_name, used_names)
                _write_dataframe(writer, province_frame.reset_index(drop=True), sheet_name)
        else:
            result_sheet = _safe_sheet_name("排班汇总", used_names)
            _write_dataframe(writer, scheduled, result_sheet)

        anomalies = result.anomalies
        anomaly_sheet = _safe_sheet_name("异常待处理", used_names)
        if anomalies.empty:
            anomalies = pd.DataFrame([{"处理状态": "无异常", "异常原因": "本次排班未发现异常"}])
        _write_dataframe(writer, anomalies, anomaly_sheet)

    return output.getvalue()
