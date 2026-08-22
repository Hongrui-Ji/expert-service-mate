from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from auto_scheduler.engine import ScheduleRunResult
from auto_scheduler.excel_io import build_output_workbook, read_input_workbook


def test_output_contains_summary_province_and_anomaly_sheets():
    scheduled = pd.DataFrame(
        [
            {
                "门店ID": "1",
                "门店名称": "=恶意公式",
                "省份": "测试省",
                "开始日期": date(2026, 8, 1),
                "截止日期": date(2026, 8, 1),
                "服务锚点日期": date(2026, 8, 1),
            }
        ]
    )
    anomalies = pd.DataFrame([{"异常类型": "测试异常", "异常原因": "示例"}])
    result = ScheduleRunResult(
        scheduled=scheduled,
        anomalies=anomalies,
        summary={"目标月份": "2026-08", "成功排班数": 1},
    )

    content = build_output_workbook(result, split_by_province=True)
    workbook = load_workbook(BytesIO(content), data_only=False)

    assert workbook.sheetnames == ["运行摘要", "测试省", "异常待处理"]
    assert workbook["测试省"]["B2"].value == "'=恶意公式"
    assert workbook["测试省"]["D2"].number_format == "yyyy/mm/dd"


def test_input_workbook_is_read_from_memory():
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame([{"门店ID": "1", "门店名称": "测试门店"}]).to_excel(writer, index=False)

    dataframe = read_input_workbook(buffer.getvalue())

    assert dataframe.to_dict("records") == [{"门店ID": 1, "门店名称": "测试门店"}]
